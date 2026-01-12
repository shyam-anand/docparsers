from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from typing import Any, Iterable

from extractors.core.extractor.base_extractor import BaseExtractor
from extractors.core.models.document_unit import DocumentUnit
import pandas as pd
from io import BytesIO
import openpyxl  # type: ignore


class ExcelExtractor(BaseExtractor):
    parser_name = "openpyxl+pandas"
    parser_version = "1.0"

    def _extract_text(self, sheet: Worksheet) -> str:
        """Extract text (best-effort, visual order)"""
        text_lines = []
        for row in sheet.iter_rows():
            values = [str(cell.value) for cell in row if cell.value is not None]
            if values:
                text_lines.append(" | ".join(values))
        return "\n".join(text_lines)

    def _extract_tables(self, sheet: Worksheet) -> list[dict[str, Any]]:
        """Extract tables (simple heuristic: entire sheet as one table)"""
        df = pd.DataFrame(sheet.values)
        tables = []
        if not df.empty:
            tables.append(
                {
                    "table_id": None,
                    "rows": df.fillna("").astype(str).values.tolist(),
                    "columns": df.iloc[0].astype(str).tolist() if len(df) > 0 else [],
                }
            )
        return tables

    def extract(self, document_id: str, file_bytes: bytes) -> Iterable[DocumentUnit]:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        for idx, sheet_name in enumerate[str](workbook.sheetnames):
            sheet = workbook[sheet_name]

            raw_text = self._extract_text(sheet)
            tables = self._extract_tables(sheet)
            layout = {
                "merged_cells": bool(sheet.merged_cells),
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }

            yield DocumentUnit(
                document_id=document_id,
                unit_type="sheet",
                unit_index=idx,
                unit_name=sheet_name,
                raw_text=raw_text,
                tables=tables,
                layout=layout,
                parser_name=self.parser_name,
                parser_version=self.parser_version,
            )
