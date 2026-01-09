from extractors.core.extractor.base_extractor import BaseExtractor
from extractors.core.models.document_unit import DocumentUnit
from typing import List

import pandas as pd
from io import BytesIO
import openpyxl


class ExcelExtractor(BaseExtractor):
    parser_name = "openpyxl+pandas"
    parser_version = "1.0"

    def extract(self, document_id: str, file_bytes: bytes) -> List[DocumentUnit]:
        units = []

        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # Extract text (best-effort, visual order)
            text_lines = []
            for row in ws.iter_rows():
                values = [str(cell.value) for cell in row if cell.value is not None]
                if values:
                    text_lines.append(" | ".join(values))
            raw_text = "\n".join(text_lines)

            # Extract tables (simple heuristic: entire sheet as one table)
            df = pd.DataFrame(ws.values)
            tables = []
            if not df.empty:
                tables.append(
                    {
                        "table_id": None,
                        "rows": df.fillna("").astype(str).values.tolist(),
                        "columns": df.iloc[0].astype(str).tolist() if len(df) > 0 else [],
                    }
                )

            layout = {
                "merged_cells": bool(ws.merged_cells),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }

            units.append(
                DocumentUnit(
                    document_id=document_id,
                    unit_type="sheet",
                    unit_index=idx,
                    unit_name=sheet_name,
                    raw_text=raw_text,
                    tables=tables,
                    layout=layout,
                    parser_name=self.parser_name,
                    parser_version=self.parser_version,
                )
            )

        return units
